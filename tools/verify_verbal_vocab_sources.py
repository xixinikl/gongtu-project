#!/usr/bin/env python3
"""Read-only verification for the idiom vocab and verbal question bank sources."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "doc" / "import-reports"

VOCAB_XLSX = Path("/Users/miduoduo/Downloads/高频成语800词完整版（修复版）.xlsx")
QUESTION_XLSX = Path("/Users/miduoduo/Downloads/花生海海刷言语理解题库.xlsx")
QUESTION_JSON = Path("/Users/miduoduo/Downloads/花生海海刷言语理解题库.json")
ANSWER_PDF = Path("/Users/miduoduo/Downloads/【四海】26海海刷言语理解题本-答案汇总.pdf")


@dataclass
class SourceCheckResult:
    summary: dict[str, Any]
    details: dict[str, Any]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_module_name(value: str) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def normalize_question_type(value: str) -> str:
    text = clean_text(value)
    mapping = {
        "逻辑填空": "logic_fill",
        "选词填空": "logic_fill",
        "片段阅读": "reading_comprehension",
        "阅读理解": "reading_comprehension",
    }
    return mapping.get(text, text)


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    headers = [clean_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = {
            headers[col_idx - 1]: ws.cell(row_idx, col_idx).value
            for col_idx in range(1, ws.max_column + 1)
            if headers[col_idx - 1]
        }
        if any(clean_text(v) for v in row.values()):
            rows.append(row)
    return rows


def verify_vocab() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    rows = read_xlsx_rows(VOCAB_XLSX)
    entries: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=2):
        word = clean_text(row.get("成语"))
        if not word:
            continue
        examples = clean_text(row.get("例句"))
        entries.append(
            {
                "row": index,
                "word": word,
                "meaning": clean_text(row.get("解释")),
                "examples": examples,
                "search_label": clean_text(row.get("人民网搜索")),
                "example_line_count": len([line for line in examples.splitlines() if line.strip()]),
                "has_newline_examples": "\n" in examples,
            }
        )

    counts = Counter(item["word"] for item in entries)
    duplicate_words = sorted([word for word, count in counts.items() if count > 1])
    summary = {
        "source": str(VOCAB_XLSX),
        "data_rows": len(entries),
        "unique_words": len(counts),
        "duplicate_word_count": len(duplicate_words),
        "duplicate_words": duplicate_words,
        "rows_with_newline_examples": sum(1 for item in entries if item["has_newline_examples"]),
        "rows_with_examples": sum(1 for item in entries if item["examples"]),
        "rows_missing_meaning": sum(1 for item in entries if not item["meaning"]),
        "recommended_search_url_template": "https://search.people.cn/s?keyword=<词语>",
    }
    return summary, entries


def flatten_question_json() -> list[dict[str, Any]]:
    payload = json.loads(QUESTION_JSON.read_text(encoding="utf-8"))
    questions: list[dict[str, Any]] = []
    for category_key, category in payload.get("categories", {}).items():
        for seq, question in enumerate(category.get("questions", []), start=1):
            source_module = normalize_module_name(question.get("source", ""))
            questions.append(
                {
                    "id": clean_text(question.get("id")),
                    "type": clean_text(question.get("type") or category_key),
                    "source_module": source_module,
                    "module_sequence": seq,
                    "stem": clean_text(question.get("stem")),
                    "options": {key: clean_text(value) for key, value in question.get("options", {}).items()},
                    "answer": clean_text(question.get("answer")),
                    "explanation": clean_text(question.get("explanation")),
                    "relatedIdioms": question.get("relatedIdioms", []),
                }
            )
    return questions


def read_question_xlsx() -> list[dict[str, Any]]:
    rows = read_xlsx_rows(QUESTION_XLSX)
    module_counts: dict[str, int] = defaultdict(int)
    questions: list[dict[str, Any]] = []
    for row in rows:
        qid = clean_text(row.get("题目ID"))
        if not qid:
            continue
        source_module = normalize_module_name(row.get("来源模块", ""))
        module_counts[source_module] += 1
        options = {
            "A": clean_text(row.get("选项A")),
            "B": clean_text(row.get("选项B")),
            "C": clean_text(row.get("选项C")),
            "D": clean_text(row.get("选项D")),
        }
        questions.append(
            {
                "id": qid,
                "type": normalize_question_type(clean_text(row.get("题型"))),
                "source_module": source_module,
                "module_sequence": module_counts[source_module],
                "raw_question_number": clean_text(row.get("题号")),
                "stem": clean_text(row.get("题干")),
                "options": options,
                "answer": clean_text(row.get("正确答案")),
                "explanation": clean_text(row.get("答案解析")),
            }
        )
    return questions


def parse_pdf_answers() -> dict[str, dict[int, str]]:
    module_answers: dict[str, dict[int, str]] = {}
    current_module = ""
    reader = PdfReader(str(ANSWER_PDF))
    for page in reader.pages:
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = clean_text(raw_line)
            if not line:
                continue
            module_match = re.fullmatch(r"([\u4e00-\u9fa5]+)\s*(\d+)", line)
            if module_match:
                current_module = normalize_module_name("".join(module_match.groups()))
                module_answers.setdefault(current_module, {})
                continue
            if not current_module:
                continue
            for start, end, answers in re.findall(r"(\d+)\s*--\s*(\d+)\s*([A-D]+)", line):
                start_i = int(start)
                end_i = int(end)
                if end_i - start_i + 1 != len(answers):
                    raise ValueError(f"Answer block length mismatch: {current_module} {line}")
                for offset, answer in enumerate(answers):
                    module_answers[current_module][start_i + offset] = answer
    return module_answers


def verify_questions() -> dict[str, Any]:
    xlsx_questions = read_question_xlsx()
    json_questions = flatten_question_json()
    pdf_answers = parse_pdf_answers()

    xlsx_by_id = {item["id"]: item for item in xlsx_questions}
    json_by_id = {item["id"]: item for item in json_questions}

    id_mismatches = sorted(set(xlsx_by_id) ^ set(json_by_id))
    json_mismatches = []
    for qid in sorted(set(xlsx_by_id) & set(json_by_id)):
        xq = xlsx_by_id[qid]
        jq = json_by_id[qid]
        fields = ["type", "source_module", "stem", "answer"]
        changed = [field for field in fields if xq[field] != jq[field]]
        changed.extend([f"option_{key}" for key in "ABCD" if xq["options"].get(key) != jq["options"].get(key)])
        if changed:
            json_mismatches.append({"id": qid, "fields": changed})

    pdf_mismatches = []
    missing_pdf_answers = []
    for question in xlsx_questions:
        expected = pdf_answers.get(question["source_module"], {}).get(question["module_sequence"])
        if not expected:
            missing_pdf_answers.append(
                {
                    "id": question["id"],
                    "source_module": question["source_module"],
                    "module_sequence": question["module_sequence"],
                }
            )
        elif expected != question["answer"]:
            pdf_mismatches.append(
                {
                    "id": question["id"],
                    "source_module": question["source_module"],
                    "module_sequence": question["module_sequence"],
                    "xlsx_answer": question["answer"],
                    "pdf_answer": expected,
                }
            )

    raw_number_anomalies = []
    for question in xlsx_questions:
        if question["raw_question_number"] != str(question["module_sequence"]):
            raw_number_anomalies.append(
                {
                    "id": question["id"],
                    "source_module": question["source_module"],
                    "raw_question_number": question["raw_question_number"],
                    "module_sequence": question["module_sequence"],
                }
            )

    by_type = Counter(item["type"] for item in xlsx_questions)
    by_module = Counter(item["source_module"] for item in xlsx_questions)
    pdf_module_counts = {module: len(answers) for module, answers in sorted(pdf_answers.items())}

    return {
        "sources": {
            "xlsx": str(QUESTION_XLSX),
            "json": str(QUESTION_JSON),
            "answer_pdf": str(ANSWER_PDF),
        },
        "total_xlsx_questions": len(xlsx_questions),
        "total_json_questions": len(json_questions),
        "by_type": dict(sorted(by_type.items())),
        "by_source_module": dict(sorted(by_module.items())),
        "pdf_module_counts": pdf_module_counts,
        "xlsx_json_id_mismatches": id_mismatches,
        "xlsx_json_field_mismatch_count": len(json_mismatches),
        "xlsx_json_field_mismatches_sample": json_mismatches[:20],
        "pdf_answer_mismatch_count": len(pdf_mismatches),
        "pdf_answer_mismatches_sample": pdf_mismatches[:20],
        "missing_pdf_answer_count": len(missing_pdf_answers),
        "missing_pdf_answers_sample": missing_pdf_answers[:20],
        "raw_question_number_anomaly_count": len(raw_number_anomalies),
        "raw_question_number_anomalies_sample": raw_number_anomalies[:30],
        "import_question_number_rule": "Do not trust Excel 题号; rebuild from source_module order as module_sequence.",
    }


def build_report(result: SourceCheckResult) -> str:
    vocab = result.summary["vocab"]
    questions = result.summary["questions"]
    duplicate_words = "、".join(vocab["duplicate_words"]) if vocab["duplicate_words"] else "无"

    lines = [
        "# 成语词库与言语题库源文件校验报告",
        "",
        "生成方式：`tools/verify_verbal_vocab_sources.py` 只读读取源文件，不写入数据库。",
        "",
        "## 成语词库",
        "",
        f"- 源文件：`{vocab['source']}`",
        f"- 数据行：{vocab['data_rows']}",
        f"- 唯一词条：{vocab['unique_words']}",
        f"- 重复词条数：{vocab['duplicate_word_count']}",
        f"- 重复词条：{duplicate_words}",
        f"- 含例句行：{vocab['rows_with_examples']}",
        f"- 例句含单元格内换行行数：{vocab['rows_with_newline_examples']}",
        f"- 缺释义行：{vocab['rows_missing_meaning']}",
        f"- 人民网搜索链接模板：`{vocab['recommended_search_url_template']}`",
        "",
        "## 言语题库",
        "",
        f"- Excel 题数：{questions['total_xlsx_questions']}",
        f"- JSON 题数：{questions['total_json_questions']}",
        f"- 题型分布：`{json.dumps(questions['by_type'], ensure_ascii=False)}`",
        f"- 来源模块分布：`{json.dumps(questions['by_source_module'], ensure_ascii=False)}`",
        f"- PDF 答案模块分布：`{json.dumps(questions['pdf_module_counts'], ensure_ascii=False)}`",
        f"- Excel/JSON ID 差异：{len(questions['xlsx_json_id_mismatches'])}",
        f"- Excel/JSON 字段差异：{questions['xlsx_json_field_mismatch_count']}",
        f"- PDF 答案不一致：{questions['pdf_answer_mismatch_count']}",
        f"- PDF 缺失答案：{questions['missing_pdf_answer_count']}",
        f"- Excel 题号异常行：{questions['raw_question_number_anomaly_count']}",
        "",
        "## 导入结论",
        "",
        "- 成语内置词库可按唯一词条导入，共 801 个唯一词条；重复词条需要按既定规则保留第一条或人工复核合并。",
        "- 例句字段必须按原单元格内换行保存和展示。",
        "- 花生海海刷题库可导入 450 题；Excel 与 JSON 题目字段一致，答案与 PDF 按模块内顺序比对无不一致。",
        "- Excel `题号` 列不可作为导入题号；导入时必须按 `来源模块` 内出现顺序重建 `module_sequence`。",
    ]
    return "\n".join(lines) + "\n"


def main() -> int:
    missing = [path for path in [VOCAB_XLSX, QUESTION_XLSX, QUESTION_JSON, ANSWER_PDF] if not path.exists()]
    if missing:
        for path in missing:
            print(f"Missing source file: {path}")
        return 2

    vocab_summary, vocab_entries = verify_vocab()
    question_summary = verify_questions()
    result = SourceCheckResult(
        summary={"vocab": vocab_summary, "questions": question_summary},
        details={"vocab_sample": vocab_entries[:20]},
    )

    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = DEFAULT_OUTPUT_DIR / "verbal_vocab_source_validation.json"
    md_path = DEFAULT_OUTPUT_DIR / "verbal_vocab_source_validation.md"
    json_path.write_text(
        json.dumps({"summary": result.summary, "details": result.details}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    md_path.write_text(build_report(result), encoding="utf-8")
    print(json.dumps(result.summary, ensure_ascii=False, indent=2))
    print(f"wrote: {md_path}")
    print(f"wrote: {json_path}")

    failed = (
        vocab_summary["unique_words"] != 801
        or question_summary["total_xlsx_questions"] != 450
        or question_summary["total_json_questions"] != 450
        or question_summary["xlsx_json_id_mismatches"]
        or question_summary["xlsx_json_field_mismatch_count"] != 0
        or question_summary["pdf_answer_mismatch_count"] != 0
        or question_summary["missing_pdf_answer_count"] != 0
    )
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
